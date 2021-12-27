import pandas as pd
import sys
import os
import openpyxl
from openpyxl import Workbook
# Import json module
import json



#############################################
#  MAIN
#############################################
if __name__ == '__main__':

    json_data = {}
    path = r"data\name_cus.json"

    with open(path) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()
    #When you make a new workbook you get a new blank active sheet
    #We need to delete it since we do not want it
    wb.remove(wb.active)

    wb1 = wb.create_sheet(title="Data")
   
    wb1.cell(1,1, "Khách hàng")
    wb1.cell(2,1, "Mã VIN")
    wb1.cell(3,1, "Biển số")
    wb1.cell(4,1, "Số điện thoại")
    wb1.cell(5,1, "Địa chỉ")
    wb1.cell(6,1, "Ngày sửa chữa")
    wb1.cell(7,1, "Hư hỏng")
    wb1.cell(8,1, "Xilanh 1")
    wb1.cell(9,1, "Xilanh 2")
    wb1.cell(10,1, "Xilanh 3")
    wb1.cell(11,1, "Xilanh 4")
    
    wb1.cell(1,2, json_data["name"])
    wb1.cell(2,2, json_data["VIN_code"])
    wb1.cell(3,2, json_data["number_plate"])
    wb1.cell(4,2, json_data["phone_number"])
    wb1.cell(5,2, json_data["address"])
    wb1.cell(6,2, json_data["fixing_date"])
    wb1.cell(7,2, json_data["damaged"])
    wb1.cell(8,2, "Xilanh 1")
    wb1.cell(9,2, "Xilanh 2")
    wb1.cell(10,2, "Xilanh 3")
    wb1.cell(11,2, "Xilanh 4")
    
   #Save it to excel
    wb.save("output\data.xlsx")