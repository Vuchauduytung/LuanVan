from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import sys
import os
# Import json module
from modules.library.IO_support import *
from modules.library.Open_pdf import *
from modules.library.compare import *
import webbrowser
from openpyxl import Workbook

class Ui(QtWidgets.QMainWindow):
    """
    Application
    """

    def __init__(self, main_path: str, phone_number: str):
        super(Ui, self).__init__()
        self.main_path = main_path
        self.phone_number = phone_number
        # load GUI
        gui_path = os.path.abspath(
            os.path.join(main_path, "GUI", "GUIdisplay.ui"))
        uic.loadUi(gui_path, self)
        self.setup_pushButton()
        self.setup_lineEdit_information_custom()
        self.setup_lineEdit_car_number_VIN()
        self.setup_lineEdit_diagnose()
        self.icon()
        self.show()
        self.setup_action()
        self.GUIgraph_path = os.path.abspath(os.path.join(main_path, "GUIgraph.py"))

    def setup_action(self):
        actionExport: QAction = self.findChild(QAction, "actionExport")
        actionExport.triggered.connect(self.export_exel)

    def export_exel(self):
        data_path = os.path.abspath(os.path.join(self.main_path, "data", "customers_data.json"))
        customers_list = json2dict(data_path)
        for customer in customers_list:
            if customer.get('phone_number') == self.phone_number:
                break
        else:
            raise Exception("Cannot find customer with phone number \"{phone_number}\""\
                .format(phone_number=self.phone_number))
        TE_diagnose_1: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_1")
        TE_diagnose_2: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_2")
        TE_diagnose_3: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_3")
        TE_diagnose_4: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_4")
        wb = Workbook()
        #When you make a new workbook you get a new blank active sheet
        #We need to delete it since we do not want it
        wb.remove(wb.active)
        wb1 = wb.create_sheet(title="Data")
        wb1.cell(1,1, "Khách hàng")
        wb1.cell(2,1, "Mã VIN")
        wb1.cell(3,1, "Biển số")
        wb1.cell(4,1, "Số điện thoại")
        wb1.cell(5,1, "Địa chỉ")
        wb1.cell(6,1, "Ngày sửa chữa")
        wb1.cell(7,1, "Hư hỏng")
        wb1.cell(8,1, "Xilanh 1")
        wb1.cell(9,1, "Xilanh 2")
        wb1.cell(10,1, "Xilanh 3")
        wb1.cell(11,1, "Xilanh 4")
        wb1.cell(1,2, customer["name"])
        wb1.cell(2,2, customer["VIN_code"])
        wb1.cell(3,2, customer["number_plate"])
        wb1.cell(4,2, customer["phone_number"])
        wb1.cell(5,2, customer["address"])
        wb1.cell(6,2, customer["fixing_date"])
        wb1.cell(7,2, customer["damaged"])
        wb1.cell(8,2, TE_diagnose_1.toPlainText())
        wb1.cell(9,2, TE_diagnose_2.toPlainText())
        wb1.cell(10,2, TE_diagnose_3.toPlainText())
        wb1.cell(11,2, TE_diagnose_4.toPlainText())
        wb1.cell(12,2, "Cố vấn dịch vụ")
        wb1.cell(12,3, "phamthinh02@gmail.com")
        #Save it to excel
        filepath = self.get_saveFile_directory()
        wb.save(filepath)
        
    def get_saveFile_directory(self):
        default_dir = os.path.abspath(os.path.join(self.main_path, "output"))
        default_filename = os.path.join(default_dir, "customer.xls")
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save audio file", default_filename, "Excel Files (*.xls)"
        )
        if filename:
            print(filename)
        return filename
    
    def icon(self):
        icon_path = os.path.abspath(os.path.join(self.main_path, "icon","Logo BK.png"))
        self.setWindowIcon(QIcon(icon_path))
        
    def setup_lineEdit_information_custom(self):
        #Open json
        data_path = os.path.abspath(os.path.join(self.main_path, "data", "customers_data.json"))
        customers_list = json2dict(data_path)
        for customer in customers_list:
            if customer.get('phone_number') == self.phone_number:
                break
        else:
            raise Exception("Cannot find customer with phone number \"{phone_number}\""\
                .format(phone_number=self.phone_number))
        GB_information_custom: QGroupBox = self.findChild(QGroupBox, "GB_information_custom")
        LE_name: QLineEdit = self.findChild(QLineEdit, "LE_name")
        LE_num_vin: QLineEdit = self.findChild(QLineEdit, "LE_num_vin")
        LE_lices_num: QLineEdit = self.findChild(QLineEdit, "LE_lices_num")
        LE_num_phone: QLineEdit = self.findChild(QLineEdit, "LE_num_phone")
        LE_address: QLineEdit = self.findChild(QLineEdit, "LE_address")
        LE_date_fix: QLineEdit = self.findChild(QLineEdit, "LE_date_fix")
        LE_fix_car: QLineEdit = self.findChild(QLineEdit, "LE_fix_car")
        
        #Hien gia tri
        LE_name.setText(customer["name"])
        LE_num_vin.setText(customer["VIN_code"])
        LE_lices_num.setText(customer["number_plate"])
        LE_num_phone.setText(customer["phone_number"])
        LE_address.setText(customer["address"])
        LE_date_fix.setText(customer["fixing_date"])
        LE_fix_car.setText(customer["damaged"])
        
    def setup_lineEdit_car_number_VIN(self):
        #Open json
        data_path = os.path.abspath(os.path.join(self.main_path, "data", "customers_data.json"))
        customers_list = json2dict(data_path)
        for customer in customers_list:
            if customer.get('phone_number') == self.phone_number:
                break
        else:
            raise Exception("Cannot find customer with phone number \"{phone_number}\""\
                .format(phone_number=self.phone_number))
        vin_num = list(customer["VIN_code"])
        vin_num_product = "".join(vin_num[:3])
        vin_num_product_2 = "".join(vin_num[:2])
        num_product = "".join(vin_num[11:])
        
        data_vin = os.path.abspath(os.path.join(self.main_path, "library", "libary_VIN.json"))
        num_vin_data = json2dict(data_vin)
        # Doc ma vin
        LE_area: QLineEdit = self.findChild(QLineEdit, "LE_area")    
        LE_country: QLineEdit = self.findChild(QLineEdit, "LE_country")
        LE_car_model: QLineEdit = self.findChild(QLineEdit, "LE_car_model")
        LE_car_name: QLineEdit = self.findChild(QLineEdit, "LE_car_name")
        LE_sec_num: QLineEdit = self.findChild(QLineEdit, "LE_sec_num")
        LE_product_date: QLineEdit = self.findChild(QLineEdit, "LE_product_date")
        LE_factory: QLineEdit = self.findChild(QLineEdit, "LE_factory")
        LE_num_product: QLineEdit = self.findChild(QLineEdit, "LE_num_product")
        
        #Hien gia tri
        LE_area.setText(num_vin_data["contry"]["value"][vin_num[0]]["text"])
        LE_country.setText(num_vin_data["contry"]["value"][vin_num[0]]["children"][vin_num[1]]["text"])
        vin_product = num_vin_data["name_product_car"]["value"].get(vin_num_product)
        vin_product_2 = num_vin_data["name_product_car"]["value"].get(vin_num_product_2)
        if isinstance(vin_product, dict) and vin_product != None:
            if vin_num_product == vin_product.get("key"):
                LE_car_model.setText(vin_product.get("text"))
                LE_car_name.setText(vin_product.get("text"))
            elif vin_num_product_2 == vin_product_2.get("key"):
                LE_car_model.setText(vin_product_2.get("text"))
                LE_car_name.setText(vin_product_2.get("text"))
            else:   
                LE_car_model.setText("Not assigned")
                LE_car_name.setText("Not assigned")
        else:   
            LE_car_model.setText("Not assigned")
            LE_car_name.setText("Not assigned")    
        LE_sec_num.setText(vin_num[8])
        LE_product_date.setText(num_vin_data["product_date"]["value"][vin_num[6]]["text"])
        LE_factory.setText(vin_num[10])
        LE_num_product.setText(num_product)
    
    def setup_lineEdit_diagnose(self):
        data_path = os.path.abspath(os.path.join(self.main_path, "data" ,"customers_data.json"))
        customers_list = json2dict(data_path)
        for customer in customers_list:
            if customer.get('phone_number') == self.phone_number:
                break
        else:
            raise Exception("Cannot find customer with phone number \"{phone_number}\""\
                .format(phone_number=self.phone_number))
        num_xilanh = 0
        for num_xilanh in range(1,5):
            xilanh_str = "Xylanh_{}".format(num_xilanh)
            if customer[xilanh_str] == {}:
                continue
            Pmax = customer[xilanh_str]["Pmax"]
            compression_pressure = customer[xilanh_str]["compression_pressure"]
            Minimum_load_pressure = customer[xilanh_str]["Minimum_load_pressure"]
            minimum_pressure = customer[xilanh_str]["minimum_pressure"]
            P_in = customer[xilanh_str]["P_in"]
            P_out = customer[xilanh_str]["P_out"]
            P_out_st = customer[xilanh_str]["P_out_st"]
            Minimum_charge_pressure = customer[xilanh_str]["Minimum_charge_pressure"]
            P_compress_end=customer[xilanh_str]["compress_end"]
            
            value_str, damage_c_str = compare_c(compression_pressure = compression_pressure,
                                                Pmax = Pmax,
                                                minimum_pressure =minimum_pressure)
            
            value_in_str, damage_in_str = compare_in(P_in = P_in ,
                                                     Minimum_load_pressure = Minimum_load_pressure,
                                                     Minimum_charge_pressure=Minimum_charge_pressure)
            
            value_out_str, damage_out_str = compare_out(P_out= P_out ,
                                                        P_out_st=P_out_st,
                                                        P_compress_end=P_compress_end,
                                                        Minimum_charge_pressure=Minimum_charge_pressure)
            
            if value_str == "Bình thường" and value_in_str == "Bình thường":
                assess_str = """
                <h3>Trạng thái:</h3>
                <p>&ensp;{xilanh}</p>
                <p>&ensp;{value}</p>
                """\
                    .format(xilanh=xilanh_str,
                            value=value_str)
                TE_diagnoses: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_{num}"\
                    .format(num=num_xilanh))
                TE_diagnoses.setHtml(assess_str)
            elif value_str =='Hư hỏng' and value_in_str == 'Hư hỏng' and value_out_str == 'Hư hỏng':
                assess_str = """
                <h3>Trạng thái:</h3>
                <p>&ensp;{xilanh}</p>
                <p>&ensp;{value}</p>
                <p>&ensp;{damage}</p>
                <p>&ensp;{damage_in}</p>
                <p>&ensp;{damage_out}</p>
                """\
                    .format(xilanh=xilanh_str,
                            value=value_in_str,
                            damage=damage_c_str,
                            damage_in=damage_in_str,
                            damage_out=damage_out_str)
                TE_diagnoses: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_{num}"\
                    .format(num=num_xilanh))
                TE_diagnoses.setHtml(assess_str)
            elif value_str =='Hư hỏng' and value_in_str == 'Hư hỏng':
                assess_str = """
                <h3>Trạng thái:</h3>
                <p>&ensp;{xilanh}</p>
                <p>&ensp;{value}</p>
                <p>&ensp;{damage}</p>
                <p>&ensp;{damage_in}</p>
                """\
                    .format(xilanh=xilanh_str,
                            value=value_str,
                            damage=damage_c_str,
                            damage_in=damage_in_str)
                TE_diagnoses: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_{num}"\
                    .format(num=num_xilanh))
                TE_diagnoses.setHtml(assess_str)
            elif value_str =='Hư hỏng' and value_out_str == 'Hư hỏng':
                assess_str = """
                <h3>Trạng thái:</h3>
                <p>&ensp;{xilanh}</p>
                <p>&ensp;{value}</p>
                <p>&ensp;{damage}</p>
                <p>&ensp;{damage_out}</p>
                """\
                    .format(xilanh=xilanh_str,
                            value=value_str,
                            damage=damage_c_str,
                            damage_out=damage_out_str)
                TE_diagnoses: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_{num}"\
                    .format(num=num_xilanh))
                TE_diagnoses.setHtml(assess_str)
            elif value_str == 'Hư hỏng':
                assess_str = """
                <h3>Trạng thái:</h3>
                <p>&ensp;{xilanh}</p>
                <p>&ensp;{value}</p>
                <p>&ensp;{damage}</p>
                """\
                    .format(xilanh=xilanh_str,
                            value=value_str,
                            damage=damage_c_str)
                TE_diagnoses: QTextEdit = self.findChild(QTextEdit, "TE_diagnose_{num}"\
                    .format(num=num_xilanh))
                TE_diagnoses.setHtml(assess_str)
        
    def LE_focusOutEvent(UI, self: QLineEdit, event: QFocusEvent):
        if self.text() == "":
            self.action.setVisible(True)
        else:
            self.action.setVisible(False)
        QLineEdit.focusOutEvent(self, event)
        
    def setup_pushButton(self):
        # GB_informatin_custom QGroupBox
        BT_graph_fix: QPushButton = self.findChild(QPushButton, "BT_graph_fix")
        BT_graph_fix.clicked.connect(self.BT_graph_fix_click)
        
        BT_graph_1: QPushButton = self.findChild(QPushButton, "BT_graph_1")
        BT_graph_1.clicked.connect(self.BT_graph1_click)

        BT_graph_2: QPushButton = self.findChild(QPushButton, "BT_graph_2")
        BT_graph_2.clicked.connect(self.BT_graph2_click)

        BT_graph_3: QPushButton = self.findChild(QPushButton, "BT_graph_4")
        BT_graph_3.clicked.connect(self.BT_graph3_click)

        BT_graph_4: QPushButton = self.findChild(QPushButton, "BT_graph_5")
        BT_graph_4.clicked.connect(self.BT_graph4_click)
        
        BT_fix: QPushButton = self.findChild(QPushButton, "BT_fix")
        BT_fix.clicked.connect(self.BT_fix_click)
        
        BT_cancel: QPushButton = self.findChild(QPushButton, "BT_cancel")
        BT_cancel.clicked.connect(self.BT_cancel_click)
        
    def BT_graph_fix_click(self):
        file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_fix.py"))
        try:
            os.system('python3 "{python_script}" \"{phone_number}\"'\
                .format(python_script=file_data,
                        phone_number=self.phone_number))
        except:
            os.system('python "{python_script}" \"{phone_number}\"'\
                .format(python_script=file_data,
                        phone_number=self.phone_number))
        
    def BT_graph_click(self):
        try:
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_1.py"))
            os.system('python3 "{}"'.format(file_data))
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_2.py"))
            os.system('python3 "{}"'.format(file_data))
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_3.py"))
            os.system('python3 "{}"'.format(file_data))
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_4.py"))
            os.system('python3 "{}"'.format(file_data))
        except:
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_1.py"))
            os.system('python "{}"'.format(file_data))
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_2.py"))
            os.system('python "{}"'.format(file_data))
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_3.py"))
            os.system('python "{}"'.format(file_data))
            file_data = os.path.abspath(os.path.join(self.main_path, "GUIgraph_xylanh_4.py"))
            os.system('python "{}"'.format(file_data))

    def BT_graph1_click(self):
        try:
            os.system("python3 \"{python_script}\" \"{phone_number}\" 1 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
        except:
            os.system("python \"{python_script}\" \"{phone_number}\" 1 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))

    def BT_graph2_click(self):
        try:
            os.system("python3 \"{python_script}\" \"{phone_number}\" 2 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
        except:
            os.system("python \"{python_script}\" \"{phone_number}\" 2 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
    
    def BT_graph3_click(self):
        try:
            os.system("python3 \"{python_script}\" \"{phone_number}\" 3 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
        except:
            os.system("python \"{python_script}\" \"{phone_number}\" 3 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
    
    def BT_graph4_click(self):
        try:
            os.system("python3 \"{python_script}\" \"{phone_number}\" 4 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
        except:
            os.system("python \"{python_script}\" \"{phone_number}\" 4 "\
                .format(python_script=self.GUIgraph_path,
                        phone_number=self.phone_number))
            
        
    def BT_fix_click(self):
        data_path = os.path.abspath(os.path.join(self.main_path, "data", "customers_data.json"))
        customers_list = json2dict(data_path)
        customer: dict
        for customer in customers_list:
            if customer.get('phone_number') == self.phone_number:
                break
        else:
            raise Exception("Cannot find customer with phone number \"{phone_number}\""\
                .format(phone_number=self.phone_number))
        num_xilanh = 0
        for num_xilanh in range(0,4):
            num_xilanh +=1
            xilanh_str = "Xylanh_{}".format(num_xilanh)       
            Pmax = customer[xilanh_str]["Pmax"]
            compression_pressure = customer[xilanh_str]["compression_pressure"]
            Minimum_load_pressure = customer[xilanh_str]["Minimum_load_pressure"]
            minimum_pressure = customer[xilanh_str]["minimum_pressure"]
            P_in = customer[xilanh_str]["P_in"]
            P_out = customer[xilanh_str]["P_out"]
            P_out_st = customer[xilanh_str]["P_out_st"]
            Minimum_charge_pressure = customer[xilanh_str]["Minimum_charge_pressure"]
            P_compress_end=customer[xilanh_str]["compress_end"]
            value, path_open, path = open_c(main_path=self.main_path,
                                            compression_pressure = compression_pressure,
                                            Pmax = Pmax,
                                            minimum_pressure =minimum_pressure)
            value_in, path_open_in, path_in = open_in(main_path=self.main_path,
                                                      P_in = P_in ,
                                                      Minimum_load_pressure = Minimum_load_pressure,
                                                      Minimum_charge_pressure=Minimum_charge_pressure)
            value_out, path_open_out, path_out = open_out(main_path=self.main_path,
                                                         P_out= P_out ,
                                                         P_out_st=P_out_st,
                                                         P_compress_end=P_compress_end,
                                                         Minimum_charge_pressure=Minimum_charge_pressure)
            if value =='Hư hỏng' and value_in == 'Hư hỏng'and value_out == 'Hư hỏng':
                webbrowser.open_new(path_open)
                webbrowser.open_new(path)
                webbrowser.open_new(path_in)
                webbrowser.open_new(path_out)
            elif value =='Hư hỏng' and value_in == 'Hư hỏng':
                webbrowser.open_new(path_open)
                webbrowser.open_new(path)
                webbrowser.open_new(path_in)
            elif value =='Hư hỏng' and value_out == 'Hư hỏng':
                webbrowser.open_new(path_open)
                webbrowser.open_new(path)
                webbrowser.open_new(path_out)
            elif value_in == 'Hư hỏng' and value_out == 'Hư hỏng':
                webbrowser.open_new(path_open_in)
                webbrowser.open_new(path_in)
                webbrowser.open_new(path_out)
            elif value == 'Hư hỏng':
                webbrowser.open_new(path_open)
                webbrowser.open_new(path)
            elif value_in == 'Hư hỏng':
                webbrowser.open_new(path_open_in)
                webbrowser.open_new(path_in)
            elif value_out == 'Hư hỏng':
                webbrowser.open_new(path_open_out)
                webbrowser.open_new(path_out)
        
    def BT_cancel_click(self):
        self.close()
        QApplication.quit()
        
        
def main():
    main_path = os.path.abspath(os.path.dirname(__file__))
    app = QtWidgets.QApplication(sys.argv)
    if len(sys.argv) > 1:
        phone_number = sys.argv[1]
    else:
        raise Exception("Missing phone number")
    window = Ui(main_path=main_path,
                phone_number=phone_number)
    app.exec_()

if __name__ == "__main__":
    main()
